#!/usr/bin/env python3
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Optional

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


UPLOAD_DIR = Path("/home/ubuntu/.cursor/projects/workspace/uploads")
OUTPUT_FILE = Path("/workspace/banque_client_2025.xlsx")


DATE_SHORT_RE = re.compile(r"^\d{2}\.\d{2}$")
DATE_FULL_RE = re.compile(r"^\d{2}\.\d{2}\.\d{2}$")
AMOUNT_TOKEN_RE = re.compile(r"^\d+(?:,\d{2})?$")
MONTH_RE = re.compile(r"^(\d{2})")


NOISE_PREFIXES = (
    "DATE ",
    "ECRITURES",
    "TOTAUX",
    "SOLDE",
    "ANCIEN SOLDE",
    "Page",
    "Ce document",
    "Crédit Lyonnais",
    "Votre conseiller",
    "RELEVE",
    "Indicatif",
    "Prenez",
    "avec votre",
    "Internet",
    "Mobile",
    "Les sommes",
    "Garantie",
    "formulaire disponible",
)


@dataclass
class Transaction:
    date_courte: str
    date_valeur: str
    debit: str = ""
    credit: str = ""
    lib_lines: List[str] = field(default_factory=list)
    last_y: float = 0.0

    @property
    def libelle(self) -> str:
        chunks = [chunk.strip() for chunk in self.lib_lines if chunk.strip()]
        return " | ".join(chunks)


def normalize_text(value: str) -> str:
    return " ".join(value.replace("\t", " ").split()).strip()


def cluster_lines(words: List[dict], tolerance: float = 1.1) -> List[dict]:
    lines: List[dict] = []
    for word in sorted(words, key=lambda item: (item["top"], item["x0"])):
        y = float(word["top"])
        for line in lines:
            if abs(y - line["y"]) <= tolerance:
                line["words"].append(word)
                line["sum_y"] += y
                line["count"] += 1
                line["y"] = line["sum_y"] / line["count"]
                break
        else:
            lines.append({"y": y, "sum_y": y, "count": 1, "words": [word]})

    for line in lines:
        line["words"].sort(key=lambda item: item["x0"])
    lines.sort(key=lambda item: item["y"])
    return lines


def join_amount_tokens(tokens: List[str]) -> str:
    useful = [token for token in tokens if token != "."]
    return "".join(useful)


def parse_french_amount(amount: str) -> Optional[float]:
    if not amount:
        return None
    cleaned = amount.replace(" ", "").replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return None


def parse_date_value(date_value: str) -> str:
    day, month, year = date_value.split(".")
    return f"20{year}-{month}-{day}"


def is_noise_line(text: str) -> bool:
    compact = normalize_text(text)
    if not compact:
        return True
    return any(compact.startswith(prefix) for prefix in NOISE_PREFIXES)


def parse_statement(pdf_path: Path) -> List[Transaction]:
    transactions: List[Transaction] = []

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            current: Optional[Transaction] = None
            words = page.extract_words(use_text_flow=False, keep_blank_chars=False)
            lines = cluster_lines(words)

            for line in lines:
                words_in_line = line["words"]
                y = float(line["y"])

                short_dates = [
                    word["text"]
                    for word in words_in_line
                    if word["x0"] < 80 and DATE_SHORT_RE.match(word["text"])
                ]
                full_dates = [
                    word["text"]
                    for word in words_in_line
                    if 330 <= word["x0"] <= 430 and DATE_FULL_RE.match(word["text"])
                ]

                debit_tokens = [
                    word["text"]
                    for word in words_in_line
                    if 430 <= word["x0"] < 500
                    and (AMOUNT_TOKEN_RE.match(word["text"]) or word["text"] == ".")
                ]
                credit_tokens = [
                    word["text"]
                    for word in words_in_line
                    if 500 <= word["x0"] < 570
                    and (AMOUNT_TOKEN_RE.match(word["text"]) or word["text"] == ".")
                ]

                debit = join_amount_tokens(debit_tokens)
                credit = join_amount_tokens(credit_tokens)
                has_amount = bool(debit or credit)

                if short_dates and full_dates and has_amount:
                    if current:
                        transactions.append(current)

                    lib_same_line = normalize_text(
                        " ".join(
                            word["text"]
                            for word in words_in_line
                            if 70 <= word["x0"] < 340
                        )
                    )
                    current = Transaction(
                        date_courte=short_dates[0],
                        date_valeur=full_dates[0],
                        debit=debit,
                        credit=credit,
                        lib_lines=[lib_same_line] if lib_same_line else [],
                        last_y=y,
                    )
                    continue

                if not current:
                    continue

                # Les lignes de continuation d'un libellé sont proches de la ligne
                # d'écriture précédente. Au-delà, on considère que c'est du bruit de page.
                if y - current.last_y > 15:
                    continue

                lib_following_line = normalize_text(
                    " ".join(
                        word["text"] for word in words_in_line if 70 <= word["x0"] < 340
                    )
                )

                if lib_following_line and not is_noise_line(lib_following_line):
                    current.lib_lines.append(lib_following_line)
                    current.last_y = y

            if current:
                transactions.append(current)

    return transactions


def list_pdf_files() -> List[Path]:
    files = sorted(UPLOAD_DIR.glob("*.pdf"))
    return sorted(
        files,
        key=lambda path: int(MONTH_RE.search(path.name).group(1))
        if MONTH_RE.search(path.name)
        else 99,
    )


def build_workbook(statements: List[tuple[int, Path, List[Transaction]]]) -> None:
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    headers = ["Date", "JNL", "CPT", "Piece", "Lib", "D", "C"]

    for month, _, transactions in statements:
        ws = wb.create_sheet(title=f"{month:02d}")
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for tx in transactions:
            ws.append(
                [
                    parse_date_value(tx.date_valeur),
                    "BQ",
                    "",
                    "",
                    tx.libelle,
                    parse_french_amount(tx.debit),
                    parse_french_amount(tx.credit),
                ]
            )

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            row[0].number_format = "dd/mm/yyyy"
            row[5].number_format = "#,##0.00"
            row[6].number_format = "#,##0.00"

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 8
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 95
        ws.column_dimensions["F"].width = 14
        ws.column_dimensions["G"].width = 14
        ws.freeze_panes = "A2"

    wb.save(OUTPUT_FILE)


def main() -> None:
    pdf_files = list_pdf_files()
    if not pdf_files:
        raise FileNotFoundError(f"Aucun PDF trouvé dans {UPLOAD_DIR}")

    statements: List[tuple[int, Path, List[Transaction]]] = []
    for pdf_file in pdf_files:
        month_match = MONTH_RE.search(pdf_file.name)
        if not month_match:
            continue
        month = int(month_match.group(1))
        transactions = parse_statement(pdf_file)
        statements.append((month, pdf_file, transactions))

    build_workbook(statements)

    for month, pdf_file, transactions in statements:
        debit_total = sum(
            parse_french_amount(tx.debit) or 0.0 for tx in transactions if tx.debit
        )
        credit_total = sum(
            parse_french_amount(tx.credit) or 0.0 for tx in transactions if tx.credit
        )
        print(
            f"{month:02d} | {pdf_file.name} | lignes={len(transactions)} "
            f"| debit={debit_total:.2f} | credit={credit_total:.2f}"
        )
    print(f"\nClasseur créé : {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
